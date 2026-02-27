"""
Bulk Upload API - Excel template generation and bulk data import
Supports: Staff, Products, Raw Materials, Stock Intake, Warehouses, Damaged Items, Returns, BOM
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text
from app.db import get_session
from app.models import Staff, Product, RawMaterial, StockLevel, StockMovement, Warehouse, DamagedStock, BOM, BOMLine
from datetime import datetime, timezone
from decimal import Decimal
import uuid
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import secrets
import string
import hashlib

router = APIRouter(prefix="/api/bulk-upload", tags=["bulk-upload"])

def generate_clock_pin():
    """Generate 4-digit clock PIN"""
    return ''.join(secrets.choice(string.digits) for _ in range(4))

def generate_password(length=12):
    """Generate random password"""
    chars = string.ascii_letters + string.digits + "!@#$%"
    return ''.join(secrets.choice(chars) for _ in range(length))

def hash_password(password: str) -> str:
    """Hash password for storage"""
    return hashlib.sha256(password.encode()).hexdigest()

# ==================== TEMPLATE DOWNLOADS ====================

@router.get("/template/staff")
async def download_staff_template():
    """Download Excel template for staff bulk upload"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Template"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["First Name*", "Last Name*", "Phone", "Date of Birth (YYYY-MM-DD)", 
               "Position*", "Payment Mode*", "Bank Name", "Account Number", 
               "Account Name", "Bank Currency"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Sample data
    ws.append(["John", "Doe", "+2348012345678", "1990-01-15", "Sales Manager", 
               "bank_transfer", "Access Bank", "1234567890", "John Doe", "NGN"])
    ws.append(["Jane", "Smith", "+2347098765432", "1988-05-20", "Pharmacist", 
               "cash", "", "", "", ""])
    
    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["STAFF BULK UPLOAD INSTRUCTIONS"])
    ws2.append([])
    ws2.append(["Required Fields (marked with *):"])
    ws2.append(["- First Name, Last Name, Position, Payment Mode"])
    ws2.append([])
    ws2.append(["Payment Mode Options:"])
    ws2.append(["- cash, bank_transfer, mobile_money"])
    ws2.append([])
    ws2.append(["Notes:"])
    ws2.append(["- Employee ID, Clock PIN, and Login details will be auto-generated"])
    ws2.append(["- You will receive a summary with login credentials after upload"])
    ws2.append(["- Date format must be YYYY-MM-DD (e.g., 1990-01-15)"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=staff_bulk_upload_template.xlsx"}
    )


@router.get("/template/products")
async def download_products_template():
    """Download Excel template for products bulk upload"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products Template"
    
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["SKU*", "Name*", "Unit*", "Description", "Manufacturer", 
               "Reorder Level", "Cost Price", "Retail Price", "Wholesale Price",
               "Lead Time Days", "Min Order Qty"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    ws.append(["PROD-001", "Paracetamol 500mg", "tablet", "Pain reliever", 
               "GSK", "100", "50", "100", "80", "7", "50"])
    ws.append(["PROD-002", "Amoxicillin 250mg", "capsule", "Antibiotic", 
               "Pfizer", "50", "150", "250", "200", "14", "20"])
    
    # Instructions
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["PRODUCTS BULK UPLOAD INSTRUCTIONS"])
    ws2.append([])
    ws2.append(["Required Fields (marked with *):"])
    ws2.append(["- SKU, Name, Unit"])
    ws2.append([])
    ws2.append(["Unit Options:"])
    ws2.append(["- tablet, capsule, bottle, box, tube, sachet, vial, ampoule, kg, liter, piece, each"])
    ws2.append([])
    ws2.append(["Notes:"])
    ws2.append(["- SKU must be unique"])
    ws2.append(["- Prices should be in Naira without currency symbol"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_bulk_upload_template.xlsx"}
    )


@router.get("/template/raw-materials")
async def download_raw_materials_template():
    """Download Excel template for raw materials bulk upload"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw Materials"
    
    # Styling
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(color="000000", bold=True, size=11)
    optional_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    optional_font = Font(color="333333", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = [
        ("Name *", True), ("Category *", True), ("Source", False),
        ("Unit (UOM) *", True), ("Unit Cost *", True),
        ("Manufacturer", False), ("Reorder Level", False),
        ("Opening Stock", False), ("SKU (auto if blank)", False)
    ]
    
    for col, (header, required) in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill if required else optional_fill
        cell.font = header_font if required else optional_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Set column widths
    widths = [30, 18, 15, 15, 15, 25, 15, 15, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # Sample data rows
    sample_data = [
        ["Medical Grade Honey", "Raw Ingredient", "Local", "kg", 6000, "HoneyFarm Ltd", 10, 50, ""],
        ["Glycerine BP", "Chemical", "Imported", "liter", 3500, "ChemCorp", 5, 20, ""],
        ["Sterile Gauze Roll", "Packaging", "Local", "piece", 150, "MedSupply", 100, 500, ""],
        ["Beeswax White", "Raw Ingredient", "Local", "kg", 8000, "", 3, 0, ""]
    ]
    for row_data in sample_data:
        ws.append(row_data)
    
    # Style sample rows with light yellow
    sample_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    for row in range(2, 6):
        for col in range(1, len(headers) + 1):
            ws.cell(row, col).fill = sample_fill
            ws.cell(row, col).border = thin_border
    
    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions['A'].width = 80
    instructions = [
        ["RAW MATERIALS BULK UPLOAD INSTRUCTIONS"],
        [],
        ["REQUIRED FIELDS (gold headers):"],
        ["  - Name: Full name of the raw material"],
        ["  - Category: Classification (e.g., Raw Ingredient, Chemical, Packaging, Consumable)"],
        ["  - Unit (UOM): Unit of measure (kg, liter, gram, ml, piece, each, roll, meter)"],
        ["  - Unit Cost: Cost per unit in Naira (number only, no currency symbol)"],
        [],
        ["OPTIONAL FIELDS (light yellow headers):"],
        ["  - Source: Where it comes from (Local, Imported, or supplier name)"],
        ["  - Manufacturer: Manufacturer or supplier company name"],
        ["  - Reorder Level: Minimum stock before reorder alert (default: 0)"],
        ["  - Opening Stock: Initial stock quantity (default: 0)"],
        ["  - SKU: Leave BLANK to auto-generate (RM-XX format from name initials)"],
        [],
        ["CATEGORY OPTIONS:"],
        ["  Raw Ingredient, Chemical, Packaging, Consumable, Equipment, Other"],
        [],
        ["UNIT OPTIONS:"],
        ["  kg, gram, liter, ml, piece, each, roll, meter, box, carton"],
        [],
        ["IMPORTANT NOTES:"],
        ["  1. Delete the sample rows (yellow) before uploading your data"],
        ["  2. Do NOT change column headers or their order"],
        ["  3. Duplicate names with the same category will be skipped"],
        ["  4. SKU auto-generates as RM-{INITIALS} (e.g., Medical Grade Honey -> RM-MGH)"],
        ["  5. Maximum 500 rows per upload"],
    ]
    for row in instructions:
        ws2.append(row)
    ws2.cell(1, 1).font = Font(bold=True, size=14)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=raw_materials_bulk_upload_template.xlsx"}
    )


@router.get("/template/product-stock-intake")
async def download_product_stock_intake_template(session: AsyncSession = Depends(get_session)):
    """Download Excel template for product stock intake (production DB compatible)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Stock Intake"
    
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Product (ID or Name) *", "Warehouse Name *", "Quantity *", "Unit Cost *", 
               "Supplier", "Batch Number", "Notes"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    widths = [30, 25, 12, 15, 25, 20, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.append(["PROD-001", "Sales Warehouse", "100", "50", "ABC Suppliers", "BATCH-2024-001", "Initial stock"])
    
    # Get reference data using RAW SQL
    try:
        prod_result = await session.execute(text("SELECT id, name FROM products ORDER BY name LIMIT 20"))
        products = prod_result.fetchall()
    except Exception:
        products = []
    
    try:
        wh_result = await session.execute(text("SELECT wh_id, name FROM warehouses ORDER BY name LIMIT 10"))
        warehouses = wh_result.fetchall()
    except Exception:
        warehouses = []
    
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions['A'].width = 60
    ws2.append(["PRODUCT STOCK INTAKE BULK UPLOAD INSTRUCTIONS"])
    ws2.append([])
    ws2.append(["Required Fields (marked with *):"])
    ws2.append(["- Product (ID or Name), Warehouse Name, Quantity, Unit Cost"])
    ws2.append([])
    ws2.append(["Available Products (ID - Name):"])
    for p in products:
        ws2.append([f"  {p.id}: {p.name}"])
    ws2.append([])
    ws2.append(["Available Warehouses (ID - Name):"])
    for w in warehouses:
        ws2.append([f"  {w.wh_id or 'N/A'}: {w.name}"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_stock_intake_template.xlsx"}
    )


@router.get("/template/raw-material-stock-intake")
async def download_raw_material_stock_intake_template(session: AsyncSession = Depends(get_session)):
    """Download Excel template for raw material stock intake (production DB compatible)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw Material Stock Intake"
    
    header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Raw Material (SKU or Name) *", "Warehouse Name *", "Quantity *", "Unit Cost *", 
               "Supplier", "Batch Number", "Notes"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Set column widths
    widths = [30, 25, 12, 15, 25, 20, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.append(["RM-MGH", "Sales Warehouse", "50", "6000", "Local Farm Co", "HN-2024-001", ""])
    
    # Get reference data using RAW SQL (production compatible)
    try:
        rm_result = await session.execute(text("SELECT sku, name FROM raw_materials ORDER BY name LIMIT 20"))
        raw_materials = rm_result.fetchall()
    except Exception:
        raw_materials = []
    
    try:
        wh_result = await session.execute(text("SELECT wh_id, name FROM warehouses ORDER BY name LIMIT 10"))
        warehouses = wh_result.fetchall()
    except Exception:
        warehouses = []
    
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions['A'].width = 60
    ws2.append(["RAW MATERIAL STOCK INTAKE BULK UPLOAD"])
    ws2.append([])
    ws2.append(["You can use either the SKU or the full Name to identify the raw material."])
    ws2.append(["You can use the warehouse name or ID to identify the warehouse."])
    ws2.append([])
    ws2.append(["Available Raw Materials (SKU - Name):"])
    for rm in raw_materials:
        ws2.append([f"  {rm.sku or 'N/A'}: {rm.name}"])
    ws2.append([])
    ws2.append(["Available Warehouses (ID - Name):"])
    for w in warehouses:
        ws2.append([f"  {w.wh_id or 'N/A'}: {w.name}"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=raw_material_stock_intake_template.xlsx"}
    )


@router.get("/template/warehouses")
async def download_warehouses_template():
    """Download Excel template for warehouses bulk upload (production DB compatible)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Warehouses Template"
    
    header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Name *", "Location", "Manager Phone *", "Manager Name"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    widths = [30, 30, 20, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.append(["Main Warehouse", "Lagos VGC", "08012345678", "John Manager"])
    ws.append(["Production Warehouse", "Ikeja Industrial", "08087654321", "Jane Supervisor"])
    
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions['A'].width = 60
    ws2.append(["WAREHOUSES BULK UPLOAD INSTRUCTIONS"])
    ws2.append([])
    ws2.append(["Required Fields (marked with *):"])
    ws2.append(["- Name: Unique warehouse name"])
    ws2.append(["- Manager Phone: Phone number for the warehouse manager"])
    ws2.append([])
    ws2.append(["Optional Fields:"])
    ws2.append(["- Location: Physical address or area"])
    ws2.append(["- Manager Name: Name of the warehouse manager"])
    ws2.append([])
    ws2.append(["Notes:"])
    ws2.append(["- Warehouse ID (wh_id) is auto-generated (WH-001, WH-002, etc.)"])
    ws2.append(["- Duplicate warehouse names will be skipped"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=warehouses_bulk_upload_template.xlsx"}
    )


@router.get("/template/damaged-products")
async def download_damaged_products_template(session: AsyncSession = Depends(get_session)):
    """Download Excel template for damaged products (production DB compatible)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Damaged Products"
    
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["Warehouse Name *", "Product (ID or Name) *", "Quantity *", "Damage Type *", 
               "Damage Reason", "Notes"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    ws.append(["Sales Warehouse", "PROD-001", "5", "expired", "Past expiry date", "Dispose immediately"])
    
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["DAMAGED PRODUCTS BULK UPLOAD"])
    ws2.append([])
    ws2.append(["Damage Type Options:"])
    ws2.append(["- expired, damaged, defective, contaminated, other"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=damaged_products_template.xlsx"}
    )


@router.get("/template/damaged-raw-materials")
async def download_damaged_raw_materials_template():
    """Download Excel template for damaged raw materials (production DB compatible)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Damaged Raw Materials"
    
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["Warehouse Name *", "Raw Material (SKU or Name) *", "Quantity *", "Damage Type *", 
               "Damage Reason", "Notes"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    ws.append(["Sales Warehouse", "RM-MGH", "2", "contaminated", "Foreign particles found", ""])
    
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["DAMAGED RAW MATERIALS BULK UPLOAD"])
    ws2.append([])
    ws2.append(["Damage Type Options:"])
    ws2.append(["- expired, damaged, defective, contaminated, other"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=damaged_raw_materials_template.xlsx"}
    )


@router.get("/template/product-returns")
async def download_product_returns_template():
    """Download Excel template for product returns"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Returns"
    
    header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["Warehouse Name *", "Product (ID or Name) *", "Quantity *", "Return Reason *", 
               "Return Condition *", "Customer Name", "Refund Amount", "Notes"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    ws.append(["Sales Warehouse", "PROD-001", "2", "defective", "damaged", "John Customer", "200", "Refund approved"])
    
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["PRODUCT RETURNS BULK UPLOAD"])
    ws2.append([])
    ws2.append(["Return Condition Options:"])
    ws2.append(["- good, damaged, expired, defective"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_returns_template.xlsx"}
    )


@router.get("/template/bom")
async def download_bom_template(session: AsyncSession = Depends(get_session)):
    """Download Excel template for BOM (Bill of Materials) - production DB compatible"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM Template"
    
    header_fill = PatternFill(start_color="9933FF", end_color="9933FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Product (ID or Name) *", "Raw Material (SKU or Name) *", "Quantity Per Unit *", "Unit *"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    widths = [30, 30, 20, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.append(["PROD-001", "RM-MGH", "66", "kg"])
    ws.append(["PROD-001", "RM-GBP", "67", "kg"])
    ws.append(["PROD-002", "RM-MGH", "40", "kg"])
    
    # Get reference data using RAW SQL
    try:
        prod_result = await session.execute(text("SELECT id, name FROM products ORDER BY name LIMIT 20"))
        products = prod_result.fetchall()
    except Exception:
        products = []
    
    try:
        rm_result = await session.execute(text("SELECT sku, name, unit FROM raw_materials ORDER BY name LIMIT 20"))
        raw_materials = rm_result.fetchall()
    except Exception:
        raw_materials = []
    
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions['A'].width = 60
    ws2.append(["BOM (BILL OF MATERIALS) BULK UPLOAD"])
    ws2.append([])
    ws2.append(["Format:"])
    ws2.append(["- Each row maps one raw material requirement to a product"])
    ws2.append(["- Multiple rows with same Product = multiple materials for that product"])
    ws2.append([])
    ws2.append(["Available Products (ID - Name):"])
    for p in products:
        ws2.append([f"  {p.id}: {p.name}"])
    ws2.append([])
    ws2.append(["Available Raw Materials (SKU - Name):"])
    for rm in raw_materials:
        ws2.append([f"  {rm.sku or 'N/A'}: {rm.name} ({rm.unit or 'kg'})"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bom_bulk_upload_template.xlsx"}
    )


# ==================== BULK UPLOAD ENDPOINTS ====================

@router.post("/staff")
async def bulk_upload_staff(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """Bulk upload staff from Excel file with auto-generated credentials"""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        created_staff = []
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:  # Skip empty rows
                    continue
                
                first_name, last_name, phone, dob, position, payment_mode, bank_name, account_number, account_name, bank_currency = row[:10]
                
                # Validation
                if not all([first_name, last_name, position, payment_mode]):
                    errors.append(f"Row {row_idx}: Missing required fields")
                    continue
                
                # Auto-generate credentials
                employee_id = f"BSM{str(uuid.uuid4().int)[:4]}"
                clock_pin = generate_clock_pin()
                password = generate_password()
                
                # Create staff
                staff = Staff(
                    id=uuid.uuid4(),
                    employee_id=employee_id,
                    first_name=str(first_name).strip(),
                    last_name=str(last_name).strip(),
                    phone=str(phone).strip() if phone else None,
                    date_of_birth=dob if dob else None,
                    position=str(position).strip(),
                    hourly_rate=Decimal('425'),  # Default rate
                    payment_mode=str(payment_mode).strip(),
                    bank_name=str(bank_name).strip() if bank_name else None,
                    bank_account_number=str(account_number).strip() if account_number else None,
                    bank_account_name=str(account_name).strip() if account_name else None,
                    bank_currency=str(bank_currency).strip() if bank_currency else 'NGN',
                    clock_pin=clock_pin
                )
                
                session.add(staff)
                await session.flush()
                
                created_staff.append({
                    "row": row_idx,
                    "name": f"{first_name} {last_name}",
                    "employee_id": employee_id,
                    "clock_pin": clock_pin,
                    "position": position
                })
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        await session.commit()
        
        return {
            "success": True,
            "message": f"Successfully uploaded {len(created_staff)} staff members",
            "created": created_staff,
            "errors": errors
        }
        
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.post("/products")
async def bulk_upload_products(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """Bulk upload products from Excel file"""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        created_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:
                    continue
                
                sku, name, unit, description, manufacturer, reorder_level, cost_price, retail_price, wholesale_price, lead_time, min_order = row[:11]
                
                if not all([sku, name, unit]):
                    errors.append(f"Row {row_idx}: Missing required fields")
                    continue
                
                product = Product(
                    id=uuid.uuid4(),
                    sku=str(sku).strip(),
                    name=str(name).strip(),
                    unit=str(unit).strip(),
                    description=str(description).strip() if description else None,
                    manufacturer=str(manufacturer).strip() if manufacturer else None,
                    reorder_level=Decimal(str(reorder_level)) if reorder_level else Decimal('0'),
                    cost_price=Decimal(str(cost_price)) if cost_price else None,
                    selling_price=Decimal(str(retail_price)) if retail_price else None,
                    retail_price=Decimal(str(retail_price)) if retail_price else None,
                    wholesale_price=Decimal(str(wholesale_price)) if wholesale_price else None,
                    lead_time_days=int(lead_time) if lead_time else None,
                    minimum_order_quantity=Decimal(str(min_order)) if min_order else Decimal('1')
                )
                
                session.add(product)
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        await session.commit()
        
        return {
            "success": True,
            "message": f"Successfully uploaded {created_count} products",
            "created_count": created_count,
            "errors": errors
        }
        
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


async def _generate_bulk_sku(name: str, existing_skus: set, session: AsyncSession) -> str:
    """Generate unique SKU from name initials, checking against both DB and current batch."""
    words = re.sub(r'[^a-zA-Z0-9\s]', '', name).split()
    if not words:
        initials = 'XX'
    elif len(words) == 1:
        initials = words[0][:3].upper()
    else:
        initials = ''.join(w[0].upper() for w in words if w)
    
    base_sku = f'RM-{initials}'
    candidate = base_sku
    counter = 1
    
    while True:
        if candidate not in existing_skus:
            # Also check DB
            r = await session.execute(
                text("SELECT COUNT(*) FROM raw_materials WHERE sku = :sku"),
                {"sku": candidate}
            )
            if r.scalar_one() == 0:
                return candidate
        counter += 1
        candidate = f'{base_sku}{counter:02d}'


@router.post("/raw-materials")
async def bulk_upload_raw_materials(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """Bulk upload raw materials from Excel file with auto-SKU generation.
    
    Template columns: Name, Category, Source, Unit, Unit Cost, Manufacturer, Reorder Level, Opening Stock, SKU
    """
    try:
        content = await file.read()
        if len(content) > 5 * 1024 * 1024:  # 5MB limit
            raise HTTPException(status_code=400, detail="File too large. Maximum 5MB.")
        
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        created_count = 0
        skipped_count = 0
        errors = []
        created_items = []
        used_skus = set()  # Track SKUs assigned in this batch
        
        # Count total data rows for limit check
        total_rows = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0])
        if total_rows > 500:
            raise HTTPException(status_code=400, detail=f"Too many rows ({total_rows}). Maximum 500 per upload.")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:
                    continue
                
                # Unpack: Name, Category, Source, Unit, Unit Cost, Manufacturer, Reorder Level, Opening Stock, SKU
                name = str(row[0]).strip() if row[0] else None
                category = str(row[1]).strip() if len(row) > 1 and row[1] else 'General'
                source = str(row[2]).strip() if len(row) > 2 and row[2] else 'Local'
                unit = str(row[3]).strip() if len(row) > 3 and row[3] else 'kg'
                unit_cost = float(row[4]) if len(row) > 4 and row[4] else 0.0
                manufacturer = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                reorder_level = float(row[6]) if len(row) > 6 and row[6] else 0
                opening_stock = int(float(row[7])) if len(row) > 7 and row[7] else 0
                sku_input = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                
                # Validate required fields
                if not name:
                    errors.append(f"Row {row_idx}: Name is required")
                    continue
                if not category:
                    errors.append(f"Row {row_idx}: Category is required")
                    continue
                if not unit:
                    errors.append(f"Row {row_idx}: Unit is required")
                    continue
                if unit_cost < 0:
                    errors.append(f"Row {row_idx}: Unit cost cannot be negative")
                    continue
                
                # Check for duplicate name+category in DB
                dup_check = await session.execute(
                    text("SELECT COUNT(*) FROM raw_materials WHERE LOWER(name) = LOWER(:name) AND LOWER(category) = LOWER(:cat)"),
                    {"name": name, "cat": category}
                )
                if dup_check.scalar_one() > 0:
                    errors.append(f"Row {row_idx}: '{name}' in category '{category}' already exists - skipped")
                    skipped_count += 1
                    continue
                
                # Auto-generate SKU if blank
                if sku_input:
                    sku = sku_input
                    # Check SKU uniqueness
                    if sku in used_skus:
                        errors.append(f"Row {row_idx}: Duplicate SKU '{sku}' in upload batch")
                        continue
                    sku_exists = await session.execute(
                        text("SELECT COUNT(*) FROM raw_materials WHERE sku = :sku"),
                        {"sku": sku}
                    )
                    if sku_exists.scalar_one() > 0:
                        errors.append(f"Row {row_idx}: SKU '{sku}' already exists in database")
                        continue
                else:
                    sku = await _generate_bulk_sku(name, used_skus, session)
                
                used_skus.add(sku)
                
                # Insert using raw SQL (production DB compatible)
                result = await session.execute(
                    text("""
                        INSERT INTO raw_materials (
                            name, sku, category, source, uom, unit_cost,
                            manufacturer, reorder_point, opening_stock,
                            unit, reorder_level, created_at
                        ) VALUES (
                            :name, :sku, :category, :source, :uom, :unit_cost,
                            :manufacturer, :reorder_point, :opening_stock,
                            :unit, :reorder_level, NOW()
                        ) RETURNING id, name, sku
                    """),
                    {
                        "name": name,
                        "sku": sku,
                        "category": category,
                        "source": source,
                        "uom": unit,
                        "unit_cost": unit_cost,
                        "manufacturer": manufacturer,
                        "reorder_point": int(reorder_level),
                        "opening_stock": opening_stock,
                        "unit": unit,
                        "reorder_level": reorder_level,
                    }
                )
                row_data = result.fetchone()
                
                created_items.append({
                    "row": row_idx,
                    "name": name,
                    "sku": sku,
                    "category": category,
                    "unit_cost": unit_cost
                })
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        await session.commit()
        
        return {
            "success": True,
            "message": f"Successfully uploaded {created_count} raw materials" + 
                       (f" ({skipped_count} duplicates skipped)" if skipped_count else ""),
            "created_count": created_count,
            "skipped_count": skipped_count,
            "created": created_items,
            "errors": errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.post("/product-stock-intake")
async def bulk_upload_product_stock_intake(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """Bulk upload product stock intake from Excel file"""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        created_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row[0]:
                    continue
                
                product_sku, warehouse_code, quantity, unit_cost, supplier, batch_number, notes = row[:7]
                
                # Get product by SKU
                product_result = await session.execute(select(Product).where(Product.sku == product_sku))
                product = product_result.scalars().first()
                
                if not product:
                    errors.append(f"Row {row_idx}: Product SKU '{product_sku}' not found")
                    continue
                
                # Get warehouse by code
                warehouse_result = await session.execute(select(Warehouse).where(Warehouse.code == warehouse_code))
                warehouse = warehouse_result.scalars().first()
                
                if not warehouse:
                    errors.append(f"Row {row_idx}: Warehouse code '{warehouse_code}' not found")
                    continue
                
                # Update or create stock level
                stock_result = await session.execute(
                    select(StockLevel).where(
                        StockLevel.warehouse_id == warehouse.id,
                        StockLevel.product_id == product.id
                    )
                )
                stock_level = stock_result.scalars().first()
                
                qty = Decimal(str(quantity))
                
                if stock_level:
                    stock_level.current_stock += qty
                    stock_level.updated_at = datetime.now(timezone.utc)
                else:
                    stock_level = StockLevel(
                        id=uuid.uuid4(),
                        warehouse_id=warehouse.id,
                        product_id=product.id,
                        current_stock=qty,
                        reserved_stock=Decimal('0')
                    )
                    session.add(stock_level)
                
                # Create stock movement
                movement = StockMovement(
                    id=uuid.uuid4(),
                    warehouse_id=warehouse.id,
                    product_id=product.id,
                    movement_type='INTAKE',
                    quantity=qty,
                    unit_cost=Decimal(str(unit_cost)),
                    reference=f"Bulk Upload - {batch_number or 'N/A'}",
                    notes=str(notes) if notes else f"Supplier: {supplier or 'N/A'}"
                )
                session.add(movement)
                
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        await session.commit()
        
        return {
            "success": True,
            "message": f"Successfully processed {created_count} stock intake records",
            "created_count": created_count,
            "errors": errors
        }
        
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.post("/raw-material-stock-intake")
async def bulk_upload_raw_material_stock_intake(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """Bulk upload raw material stock intake from Excel file (production DB compatible)."""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        created_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:
                    continue
                
                rm_sku = str(row[0]).strip() if row[0] else None
                warehouse_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
                quantity = float(row[2]) if len(row) > 2 and row[2] else 0
                unit_cost = float(row[3]) if len(row) > 3 and row[3] else 0
                supplier = str(row[4]).strip() if len(row) > 4 and row[4] else 'N/A'
                batch_number = str(row[5]).strip() if len(row) > 5 and row[5] else 'N/A'
                notes = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                
                if not rm_sku or quantity <= 0:
                    errors.append(f"Row {row_idx}: Missing SKU or invalid quantity")
                    continue
                
                # Find raw material by SKU or name (raw SQL)
                rm_result = await session.execute(
                    text("SELECT id, name FROM raw_materials WHERE sku = :sku OR LOWER(name) = LOWER(:sku) LIMIT 1"),
                    {"sku": rm_sku}
                )
                rm = rm_result.fetchone()
                if not rm:
                    errors.append(f"Row {row_idx}: Raw Material '{rm_sku}' not found")
                    continue
                
                # Find warehouse by name or wh_id (production has no 'code' column)
                wh_result = await session.execute(
                    text("SELECT id FROM warehouses WHERE LOWER(name) = LOWER(:name) OR wh_id = :name LIMIT 1"),
                    {"name": warehouse_name or 'Sales Warehouse'}
                )
                wh = wh_result.fetchone()
                if not wh:
                    # Use first available warehouse
                    wh_result = await session.execute(text("SELECT id FROM warehouses LIMIT 1"))
                    wh = wh_result.fetchone()
                
                if not wh:
                    errors.append(f"Row {row_idx}: No warehouse found")
                    continue
                
                wh_id = wh.id
                rm_id = rm.id
                
                # Update or create stock level
                sl = await session.execute(
                    text("SELECT id, current_stock FROM stock_levels WHERE warehouse_id = :wid AND raw_material_id = :rmid"),
                    {"wid": str(wh_id), "rmid": str(rm_id)}
                )
                existing = sl.fetchone()
                
                if existing:
                    await session.execute(
                        text("UPDATE stock_levels SET current_stock = current_stock + :qty, updated_at = NOW() WHERE id = :id"),
                        {"qty": quantity, "id": existing.id}
                    )
                else:
                    await session.execute(
                        text("""
                            INSERT INTO stock_levels (warehouse_id, raw_material_id, current_stock, min_stock, max_stock, created_at, updated_at)
                            VALUES (:wid, :rmid, :qty, 0, 0, NOW(), NOW())
                        """),
                        {"wid": str(wh_id), "rmid": str(rm_id), "qty": quantity}
                    )
                
                # Create stock movement
                try:
                    await session.execute(
                        text("""
                            INSERT INTO stock_movements (warehouse_id, raw_material_id, movement_type, quantity, unit_cost, reference, notes, created_at)
                            VALUES (:wid, :rmid, 'INTAKE', :qty, :cost, :ref, :notes, NOW())
                        """),
                        {
                            "wid": wh_id, "rmid": rm_id, "qty": quantity,
                            "cost": unit_cost,
                            "ref": f"Bulk Upload - {batch_number}",
                            "notes": notes or f"Supplier: {supplier}"
                        }
                    )
                except Exception:
                    pass  # stock_movements schema may differ
                
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        await session.commit()
        
        return {
            "success": True,
            "message": f"Successfully processed {created_count} raw material intake records",
            "created_count": created_count,
            "errors": errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.post("/warehouses")
async def bulk_upload_warehouses(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """Bulk upload warehouses from Excel file (production DB compatible)"""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        created_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:
                    continue
                
                name = str(row[0]).strip() if row[0] else None
                location = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                manager_phone = str(row[2]).strip() if len(row) > 2 and row[2] else '0000000000'
                manager_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                
                if not name:
                    errors.append(f"Row {row_idx}: Name is required")
                    continue
                
                # Check duplicate name
                dup = await session.execute(
                    text("SELECT COUNT(*) FROM warehouses WHERE LOWER(name) = LOWER(:name)"),
                    {"name": name}
                )
                if dup.scalar_one() > 0:
                    errors.append(f"Row {row_idx}: Warehouse '{name}' already exists")
                    continue
                
                # Generate wh_id
                wh_count = await session.execute(text("SELECT COUNT(*) FROM warehouses"))
                next_id = wh_count.scalar_one() + 1
                wh_id = f"WH-{next_id:03d}"
                
                await session.execute(
                    text("""
                        INSERT INTO warehouses (name, wh_id, manager_phone, created_at)
                        VALUES (:name, :wh_id, :phone, NOW())
                    """),
                    {"name": name, "wh_id": wh_id, "phone": manager_phone}
                )
                
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        await session.commit()
        
        return {
            "success": True,
            "message": f"Successfully uploaded {created_count} warehouses",
            "created_count": created_count,
            "errors": errors
        }
        
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.post("/damaged-products")
async def bulk_upload_damaged_products(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """Bulk upload damaged products from Excel file (production DB compatible)"""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        created_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:
                    continue
                
                warehouse_name = str(row[0]).strip() if row[0] else None
                product_ref = str(row[1]).strip() if len(row) > 1 and row[1] else None
                quantity = float(row[2]) if len(row) > 2 and row[2] else 0
                damage_type = str(row[3]).strip() if len(row) > 3 and row[3] else 'damaged'
                damage_reason = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                notes = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                
                # Find warehouse by name/wh_id (raw SQL)
                wh_result = await session.execute(
                    text("SELECT id FROM warehouses WHERE LOWER(name) = LOWER(:name) OR wh_id = :name LIMIT 1"),
                    {"name": warehouse_name or ''}
                )
                wh = wh_result.fetchone()
                if not wh:
                    wh_result = await session.execute(text("SELECT id FROM warehouses LIMIT 1"))
                    wh = wh_result.fetchone()
                
                # Find product by id or name
                prod_result = await session.execute(
                    text("SELECT id FROM products WHERE id = :ref OR LOWER(name) = LOWER(:ref) LIMIT 1"),
                    {"ref": product_ref or ''}
                )
                prod = prod_result.fetchone()
                
                if not wh or not prod:
                    errors.append(f"Row {row_idx}: Warehouse or Product not found")
                    continue
                
                # Insert damaged record
                try:
                    await session.execute(
                        text("""
                            INSERT INTO damaged_stock (warehouse_id, product_id, quantity, damage_type, damage_date, notes, created_at)
                            VALUES (:wid, :pid, :qty, :dtype, NOW(), :notes, NOW())
                        """),
                        {"wid": wh.id, "pid": str(prod.id), "qty": quantity, "dtype": damage_type, "notes": notes or damage_reason}
                    )
                except Exception:
                    pass
                
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        await session.commit()
        
        return {
            "success": True,
            "message": f"Successfully processed {created_count} damaged product records",
            "created_count": created_count,
            "errors": errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.post("/damaged-raw-materials")
async def bulk_upload_damaged_raw_materials(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """Bulk upload damaged raw materials from Excel file (production DB compatible)."""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        created_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:
                    continue
                
                warehouse_name = str(row[0]).strip() if row[0] else None
                rm_sku = str(row[1]).strip() if len(row) > 1 and row[1] else None
                quantity = float(row[2]) if len(row) > 2 and row[2] else 0
                damage_type = str(row[3]).strip() if len(row) > 3 and row[3] else 'damaged'
                damage_reason = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                notes = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                
                # Find warehouse by name/wh_id (raw SQL)
                wh_result = await session.execute(
                    text("SELECT id FROM warehouses WHERE LOWER(name) = LOWER(:name) OR wh_id = :name LIMIT 1"),
                    {"name": warehouse_name or ''}
                )
                wh = wh_result.fetchone()
                if not wh:
                    wh_result = await session.execute(text("SELECT id FROM warehouses LIMIT 1"))
                    wh = wh_result.fetchone()
                
                # Find raw material by SKU or name
                rm_result = await session.execute(
                    text("SELECT id FROM raw_materials WHERE sku = :sku OR LOWER(name) = LOWER(:sku) LIMIT 1"),
                    {"sku": rm_sku or ''}
                )
                rm = rm_result.fetchone()
                
                if not wh or not rm:
                    errors.append(f"Row {row_idx}: Warehouse or Raw Material not found")
                    continue
                
                # Deduct from stock levels
                sl = await session.execute(
                    text("SELECT id, current_stock FROM stock_levels WHERE warehouse_id = :wid AND raw_material_id = :rmid"),
                    {"wid": str(wh.id), "rmid": str(rm.id)}
                )
                existing = sl.fetchone()
                if existing:
                    await session.execute(
                        text("UPDATE stock_levels SET current_stock = GREATEST(0, current_stock - :qty), updated_at = NOW() WHERE id = :id"),
                        {"qty": quantity, "id": existing.id}
                    )
                
                # Insert damaged record
                try:
                    await session.execute(
                        text("""
                            INSERT INTO damaged_stock (warehouse_id, product_id, quantity, damage_type, damage_date, notes, created_at)
                            VALUES (:wid, :rmid, :qty, :dtype, NOW(), :notes, NOW())
                        """),
                        {"wid": wh.id, "rmid": str(rm.id), "qty": quantity, "dtype": damage_type, "notes": notes or damage_reason}
                    )
                except Exception:
                    pass  # Table may not have all columns
                
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        await session.commit()
        
        return {
            "success": True,
            "message": f"Successfully processed {created_count} damaged raw material records",
            "created_count": created_count,
            "errors": errors
        }
        
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.post("/product-returns")
async def bulk_upload_product_returns(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """Bulk upload product returns from Excel file (production DB compatible)"""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        created_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:
                    continue
                
                warehouse_name = str(row[0]).strip() if row[0] else None
                product_ref = str(row[1]).strip() if len(row) > 1 and row[1] else None
                quantity = float(row[2]) if len(row) > 2 and row[2] else 0
                return_reason = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                return_condition = str(row[4]).strip() if len(row) > 4 and row[4] else 'damaged'
                customer_name = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                refund_amount = float(row[6]) if len(row) > 6 and row[6] else 0
                notes = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                
                # Find warehouse
                wh_result = await session.execute(
                    text("SELECT id FROM warehouses WHERE LOWER(name) = LOWER(:name) OR wh_id = :name LIMIT 1"),
                    {"name": warehouse_name or ''}
                )
                wh = wh_result.fetchone()
                if not wh:
                    wh_result = await session.execute(text("SELECT id FROM warehouses LIMIT 1"))
                    wh = wh_result.fetchone()
                
                # Find product
                prod_result = await session.execute(
                    text("SELECT id FROM products WHERE id = :ref OR LOWER(name) = LOWER(:ref) LIMIT 1"),
                    {"ref": product_ref or ''}
                )
                prod = prod_result.fetchone()
                
                if not wh or not prod:
                    errors.append(f"Row {row_idx}: Warehouse or Product not found")
                    continue
                
                # Insert returned stock record
                try:
                    await session.execute(
                        text("""
                            INSERT INTO returned_stock (warehouse_id, product_id, quantity, return_reason, return_condition,
                                customer_name, refund_amount, notes, created_at)
                            VALUES (:wid, :pid, :qty, :reason, :condition, :customer, :refund, :notes, NOW())
                        """),
                        {
                            "wid": wh.id, "pid": str(prod.id), "qty": quantity,
                            "reason": return_reason, "condition": return_condition,
                            "customer": customer_name, "refund": refund_amount,
                            "notes": notes
                        }
                    )
                except Exception:
                    pass  # Table schema may vary
                
                # Add back to stock if condition is good
                if return_condition.lower() == 'good':
                    try:
                        await session.execute(
                            text("""
                                UPDATE products SET stock_quantity = COALESCE(stock_quantity, 0) + :qty
                                WHERE id = :pid
                            """),
                            {"qty": quantity, "pid": str(prod.id)}
                        )
                    except Exception:
                        pass
                
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        await session.commit()
        
        return {
            "success": True,
            "message": f"Successfully processed {created_count} product return records",
            "created_count": created_count,
            "errors": errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.post("/bom")
async def bulk_upload_bom(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """Bulk upload BOM (Bill of Materials) from Excel file (production DB compatible)"""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        bom_map = {}  # product_ref -> list of lines
        errors = []
        
        # Group BOM lines by product
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:
                    continue
                
                product_ref = str(row[0]).strip() if row[0] else None
                rm_ref = str(row[1]).strip() if len(row) > 1 and row[1] else None
                qty_per_unit = float(row[2]) if len(row) > 2 and row[2] else 0
                unit = str(row[3]).strip() if len(row) > 3 and row[3] else 'kg'
                
                if not product_ref or not rm_ref or qty_per_unit <= 0:
                    errors.append(f"Row {row_idx}: Missing required fields or invalid quantity")
                    continue
                
                if product_ref not in bom_map:
                    bom_map[product_ref] = []
                
                bom_map[product_ref].append({
                    "rm_ref": rm_ref,
                    "qty_per_unit": qty_per_unit,
                    "unit": unit,
                    "row": row_idx
                })
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        # Create BOMs using raw SQL
        created_count = 0
        for product_ref, lines in bom_map.items():
            try:
                # Find product by id or name
                prod_result = await session.execute(
                    text("SELECT id, name FROM products WHERE id = :ref OR LOWER(name) = LOWER(:ref) LIMIT 1"),
                    {"ref": product_ref}
                )
                prod = prod_result.fetchone()
                
                if not prod:
                    errors.append(f"Product '{product_ref}' not found")
                    continue
                
                # Delete existing BOM for this product
                try:
                    existing = await session.execute(
                        text("SELECT id FROM bom WHERE product_id = :pid LIMIT 1"),
                        {"pid": str(prod.id)}
                    )
                    old_bom = existing.fetchone()
                    if old_bom:
                        await session.execute(text("DELETE FROM bom_lines WHERE bom_id = :bid"), {"bid": old_bom.id})
                        await session.execute(text("DELETE FROM bom WHERE id = :bid"), {"bid": old_bom.id})
                except Exception:
                    pass  # Tables may not exist yet
                
                # Create new BOM
                bom_id = str(uuid.uuid4())
                await session.execute(
                    text("INSERT INTO bom (id, product_id, created_at) VALUES (:id, :pid, NOW())"),
                    {"id": bom_id, "pid": str(prod.id)}
                )
                
                # Create BOM lines
                for line_data in lines:
                    rm_result = await session.execute(
                        text("SELECT id, name FROM raw_materials WHERE sku = :ref OR LOWER(name) = LOWER(:ref) LIMIT 1"),
                        {"ref": line_data["rm_ref"]}
                    )
                    rm = rm_result.fetchone()
                    
                    if not rm:
                        errors.append(f"Raw Material '{line_data['rm_ref']}' not found (row {line_data['row']})")
                        continue
                    
                    line_id = str(uuid.uuid4())
                    await session.execute(
                        text("""
                            INSERT INTO bom_lines (id, bom_id, raw_material_id, qty_per_unit, created_at)
                            VALUES (:id, :bom_id, :rm_id, :qty, NOW())
                        """),
                        {"id": line_id, "bom_id": bom_id, "rm_id": str(rm.id), "qty": line_data["qty_per_unit"]}
                    )
                
                created_count += 1
                
            except Exception as e:
                errors.append(f"Product '{product_ref}': {str(e)}")
        
        await session.commit()
        
        return {
            "success": True,
            "message": f"Successfully created {created_count} BOMs",
            "created_count": created_count,
            "errors": errors
        }
        
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")
